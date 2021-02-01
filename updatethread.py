import os
import time
from datetime import datetime
from threading import Thread

import openpyxl
from PyQt5 import QtCore
from PyQt5.QtWidgets import QMessageBox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break, ColBreak, RowBreak

SLEEP = 1


class UpdatingThread(QtCore.QThread):
    show_msg_box = QtCore.pyqtSignal(str, bool)
    set_state_internal = QtCore.pyqtSignal(int)
    set_state_external = QtCore.pyqtSignal(int)
    set_state_school = QtCore.pyqtSignal(int)

    def __init__(self, **kwargs):
        super().__init__()
        self.internal = kwargs['internal']
        self.external = kwargs['external']
        self.schools = kwargs['schools']
        self.internal_file_url = kwargs['internal_file_url']
        self.school_file_url = kwargs['school_file_url']
        self.external_file_url = kwargs['external_file_url']
        self.controller = kwargs['controller']

        self.is_error = False
        self.msg = ''
        self.time = None
        self.directory = '/전보 결과'

        # self.finished.connect(self.controller.show_msg_box)

    def run(self) -> None:
        self.save()

    def save(self):
        self.is_error = False
        self.msg = ''
        self.time = '{} {}'.format(datetime.now().hour, datetime.now().minute)
        self.directory = self.directory + '_' + self.time + '/'

        t1 = Thread(target=self.save_internal)
        t2 = Thread(target=self.save_schools)
        t3 = Thread(target=self.save_external)

        t1.start()
        t2.start()
        t3.start()

        t1.join()
        t2.join()
        t3.join()

        if not self.is_error:
            self.msg = "저장되었습니다."
        self.show_msg_box.emit(self.msg, self.is_error)

    def save_internal(self):
        try:
            dir = os.path.dirname(self.internal_file_url)
            fname, ext = os.path.splitext(os.path.basename(self.internal_file_url))

            has_macro = False

            if 'xlsm' in ext:
                has_macro = True

            wb = load_workbook(filename=self.internal_file_url, read_only=False, data_only=False, keep_vba=has_macro, keep_links=True)
            ws = wb['Sheet1']

            fontStyle = Font(size="10")
            alignment = Alignment(horizontal='center', vertical='center')
            patternFill = PatternFill(patternType='solid', fgColor=Color('FD7D82'))
            # patternFill = PatternFill(patternType='solid', fgColor=Color('C2E7FF'))

            # for i in range(0, ws.__len__()):
            ws.merge_cells('Y6:Y9')
            ws['Y6'].value = '임지지정'
            ws['Y6'].font = fontStyle
            ws['Y6'].fill = patternFill
            ws['Y6'].fill = patternFill
            ws['Y6'].alignment = alignment

            count = 0
            for i in range(0, self.internal.__len__()):

                if self.internal[i].disposed is None:
                    value = ''
                else:
                    value = self.internal[i].disposed.name

                cell = ws.cell(row=self.internal[i].id + 10, column=25, value=value)
                cell.font = fontStyle
                cell.alignment = alignment
                cell.fill = patternFill

                count += 1
                self.msleep(SLEEP)
                self.set_state_internal.emit(count)

            if not os.path.exists((dir + self.directory)):
                os.makedirs(dir + self.directory)

            wb.save(dir + self.directory + fname + '_결과' + self.time + ext)
            print("internal saved")
            wb.close()

        except Exception as e:
            print(e)
            self.is_error = True
            if self.msg:
                self.msg += ', '
            self.msg += '관내명부'
            wb.close()

    def save_external(self):
        try:
            dir = os.path.dirname(self.external_file_url)
            fname, ext = os.path.splitext(os.path.basename(self.external_file_url))

            has_macro = False

            if 'xlsm' in ext:
                has_macro = True

            wb = load_workbook(self.external_file_url, read_only=False, data_only=False, keep_vba=has_macro, keep_links=True)
            ws = wb['순위명부']

            fontStyle = Font(size="8")

            count = 0
            for i in range(0, self.external.__len__()):
                # if self.external[i].disposed is None or '미충원' in self.external[i].type:
                if self.external[i].disposed is None:
                    school_name = ''
                    school_area = ''
                else:
                    school_name = self.external[i].disposed.name
                    school_area = self.external[i].disposed.area

                cell = ws.cell(row=self.external[i].id + 3, column=11, value=school_name)
                cell = ws.cell(row=self.external[i].id + 3, column=12, value=school_area)
                cell.font = fontStyle
                count += 1
                self.msleep(SLEEP)
                self.set_state_external.emit(count)

            if not os.path.exists((dir + self.directory)):
                os.makedirs(dir + self.directory)
            wb.save(dir + self.directory + fname + '_결과' + self.time + '.xlsx')
            print("external saved")
            wb.close()

        except Exception as e:
            print(e)
            self.is_error = True
            if self.msg:
                self.msg += ', '
            self.msg += '관외명부'
            wb.close()

    def save_schools(self):
        try:
            dir = os.path.dirname(self.school_file_url)
            fname, ext = os.path.splitext(os.path.basename(self.school_file_url))

            has_macro = False

            if 'xlsm' in ext:
                has_macro = True

            wb = load_workbook(self.school_file_url, read_only=False, data_only=False, keep_vba=has_macro, keep_links=True)
            ws = wb['결충원']

            fontStyle = Font(size="8")

            count = 0
            for i in range(0, self.schools.__len__()):
                cell = ws.cell(row=i + 8, column=53, value=self.schools[i].gone)
                cell.font = fontStyle
                cell = ws.cell(row=i + 8, column=54, value=self.schools[i].inside)
                cell.font = fontStyle
                cell = ws.cell(row=i + 8, column=56, value=self.schools[i].outside)
                cell.font = fontStyle
                cell = ws.cell(row=i + 8, column=64, value=self.schools[i].term)
                cell.font = fontStyle
                count += 1
                self.msleep(SLEEP)
                self.set_state_school.emit(count)

            if not os.path.exists((dir + self.directory)):
                os.makedirs(dir + self.directory)

            wb.save(dir + self.directory + fname + '_결과' + self.time + ext)
            print("schools saved")
            wb.close()

        except Exception as e:
            print(e)
            self.is_error = True

            if self.msg:
                self.msg += ', '
            self.msg += '결충원'
            wb.close()
