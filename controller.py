import os
import sys
import urllib


from PyQt5.QtCore import QObject
from PyQt5.QtWidgets import QMessageBox, QApplication
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from loadingwidget import LoadingWidget
from model.schoolstatus import SchoolStatus
from postthread import PostingThread
from savethread import SavingThread
from savingwidget import SavingWidget
from workingfield import WorkingField
from model.teacher_external import TeacherExternal
from model.teacher_internal import TeacherInternal
from updatethread import UpdatingThread
from updatingwidget import UpdatingWidget

class StartController(QObject):

    def __init__(self) -> None:
        super().__init__()

        # 관내전출희망교사
        self.internal_list = list()
        # 학교 리스트
        self.school_list = list()
        # 관외전입자
        self.external_list = list()
        # 학교 해쉬
        self.hash_schools = dict()

        # ----------- 학교별 구분 --------- #
        # 해당교 전입교사
        self.designation = list()
        # 해당교 관내전출교사
        self.gone = list()
        # 해당교에서 관외로 전출간 교사
        self.gone_external = list()
        # 해당교에서 발령가지 못한 교사
        self.unDeployed = list()
        # 해당교 결원교사
        self.vacancy = list()
        # 해당교 충원교사
        self.recruit = list()
        # ----------- 학교별 구분 --------- #

        self.internal_file_url = None
        self.external_file_url = None
        self.school_file_url = None

        self.post_thread = PostingThread(internal=self.internal_list,
                                         external=self.external_list,
                                         schools=self.school_list,
                                         hash_schools=self.hash_schools)

        self.save_thread = SavingThread(internal=self.internal_list,
                                        external=self.external_list,
                                        schools=self.school_list,
                                        hash_schools=self.hash_schools,
                                        designation=self.designation,
                                        gone=self.gone,
                                        gone_external=self.gone_external,
                                        unDeployed=self.unDeployed,
                                        controller=self)

        self.update_thread = UpdatingThread(internal=self.internal_list,
                                            external=self.external_list,
                                            schools=self.school_list,
                                            hash_schools=self.hash_schools,
                                            internal_file_url=self.internal_file_url,
                                            external_file_url=self.external_file_url,
                                            school_file_url=self.school_file_url,
                                            controller=self)

        self.loading_post = LoadingWidget(**{'controller': self,
                                             'title': '기다려주세요..',
                                             'internal': self.internal_list,
                                             'school': self.school_list,
                                             'external': self.external_list,})

        self.waiting_save = SavingWidget(**{'controller': self,
                                            'title': '저장중입니다..',
                                            'internal': self.internal_list,
                                            'school': self.school_list,
                                            'external': self.external_list,
                                            'designation': self.designation,
                                            'gone': self.gone,
                                            'gone_external': self.gone_external,
                                            'unDeployed': self.unDeployed})

        self.waiting_update = UpdatingWidget(**{'controller': self,
                                                'title': '저장중입니다..',
                                                'internal': self.internal_list,
                                                'schools': self.school_list,
                                                'external': self.external_list})

        self.init_thread()

        self.flag_internal = False
        self.flag_schools = False
        self.flag_external = False

    def init_thread(self):
        self.post_thread.started.connect(self.loading_post.show)
        self.post_thread.finished.connect(self.loading_post.close)
        self.post_thread.go_to_next.connect(self.show_next_view)

        self.update_thread.show_msg_box.connect(self.show_msg_box)
        self.update_thread.started.connect(self.waiting_update.show)
        self.update_thread.started.connect(self.waiting_update.set_maximum)
        self.update_thread.set_state_internal.connect(self.waiting_update.progressBar_internal.setValue)
        self.update_thread.set_state_external.connect(self.waiting_update.progressBar_external.setValue)
        self.update_thread.set_state_school.connect(self.waiting_update.progressBar_school.setValue)
        self.update_thread.finished.connect(self.waiting_update.close)

        self.save_thread.show_msg_box.connect(self.show_msg_box)
        self.save_thread.started.connect(self.waiting_save.show)
        self.save_thread.started.connect(self.waiting_save.set_maximum)
        self.save_thread.set_state_result.connect(self.waiting_save.progressBar_result.setValue)
        self.save_thread.finished.connect(self.waiting_save.close)

    def get_internal_list(self, file_url):
        fname, ext = os.path.splitext(file_url)

        has_macro = False
        err_code = 0

        if 'xlsm' in ext:
            has_macro = True

        try:
            wb = load_workbook(file_url, data_only=True, keep_vba=has_macro, read_only=True)

            self.internal_list.clear()
            self.internal_file_url = file_url

            err_code += 1
            ws = wb['Sheet1']
            i = 0
            for row in ws.iter_rows(min_row=10):
                if row[1].value is None:
                    break

                t = TeacherInternal(id=i, school=row[3].value, region_grade=row[8].value,
                                    name=row[4].value, sex=row[6].value, regist_num=row[5].value,
                                    type=row[18].value, transfer_year=row[16].value, transfer_score=row[17].value,
                                    current_school_year=row[10].value, first=row[19].value, second=row[20].value,
                                    third=row[21].value, date=row[9].value, remarks=row[23].value,
                                    disposed=row[24].value)

                self.internal_list.append(t)
                i += 1

            self.internal_list.sort()

        except TypeError as e:
            print(e)
            self.show_msg_box("{}시트 {}번 행 서식을 확인해주세요.".format(ws.title, row[0].row), True)

            self.flag_internal = False
            wb.close()

        except KeyError as e:
            print(e)

            if err_code == 1:
                err_sheet_name = 'Sheet1'

            self.show_msg_box("\'" + err_sheet_name + "\' 시트가 필요합니다.", True)
            self.flag_internal = False
            wb.close()

        except Exception as e:
            print(e)

            if isinstance(e, InvalidFileException):
                self.show_msg_box("파일을 선택해주세요.", True)
            else:
                self.show_msg_box("올바른 파일을 선택해주세요.", True)

            if self.internal_list.__len__() == 0:
                self.flag_internal = False

        else:
            self.flag_internal = True
            self.show_msg_box("성공적으로 불러왔습니다.", False)
            wb.close()

    def get_school_list(self, file_url):
        fname, ext = os.path.splitext(file_url)

        has_macro = False
        err_code = 0

        if 'xlsm' in ext:
            has_macro = True

        try:
            wb = load_workbook(file_url, data_only=True, keep_vba=has_macro, read_only=True)

            self.school_list.clear()
            self.vacancy.clear()
            self.recruit.clear()
            self.school_file_url = file_url

            err_code += 1
            ws = wb['결충원']

            i = 0
            for row in ws.iter_rows(min_row=8):
                if row[0].value is None:
                    break

                self.school_list.append(SchoolStatus(num=row[0].value, name=row[2].value, status=row[51].value,
                                                     outside=row[55].value, inside=row[53].value, gone=row[52].value,
                                                     term=row[63].value, area=row[1].value))

                self.hash_schools[row[2].value] = row[0].internal_value
                self.designation.append([])
                self.gone.append([])
                self.unDeployed.append([])
                self.vacancy.append([])
                self.recruit.append([])
                i += 1

            err_code += 1

            i = 0
            ws = wb['결원내용']
            for row in ws.iter_rows(min_row=4):
                if row[3].value is None:
                    break
                t = TeacherInternal(id=row[0].value, school=row[2].value, region_grade=None,
                                    position=None, name=row[3].value, sex=None, regist_num=None,
                                    type=row[4].value, transfer_year=None, transfer_score=None,
                                    first=None, second=None, third=None, current_school_year=None,
                                    date=None, remarks=None, disposed=None)

                self.vacancy[self.hash_schools.get(t.school) - 1].append(t)
                i += 1

            err_code += 1

            i = 0
            ws = wb['충원내용']
            for row in ws.iter_rows(min_row=4):
                if row[3].value is None:
                    break
                t = TeacherInternal(id=row[0].value, school=row[2].value, region_grade=None,
                                    position=None, name=row[3].value, sex=None, regist_num=None,
                                    type=row[4].value, transfer_year=None, transfer_score=None,
                                    first=None, second=None, third=None, current_school_year=None,
                                    date=None, remarks=None, disposed=None)

                self.recruit[self.hash_schools.get(t.school) - 1].append(t)
                i += 1

        except TypeError as e:
            print(e)

            self.show_msg_box("\'{}\' 시트 {}번째 행 서식을 확인해주세요.".format(ws.title, row[0].row), True)

            self.flag_schools = False
            wb.close()

        except KeyError as e:
            print(e)

            if err_code == 1:
                err_sheet_name = '결충원'
            elif err_code == 2:
                err_sheet_name = '결원내용'
            elif err_code == 3:
                err_sheet_name = '충원내용'

            self.show_msg_box("\'" + err_sheet_name + "\' 시트가 필요합니다.", True)
            self.flag_schools = False
            wb.close()

        except Exception as e:
            print(e)

            if isinstance(e, InvalidFileException):
                self.show_msg_box("파일을 선택해주세요.", True)
            else:
                self.show_msg_box("올바른 파일을 선택해주세요.", True)

            if self.school_list.__len__() == 0:
                self.flag_schools = False

        else:
            # self.designation = [[] for row in range(self.school_list.__len__())]
            # self.gone = [[] for row in range(self.school_list.__len__())]
            self.flag_schools = True
            self.show_msg_box("성공적으로 불러왔습니다.", False)
            wb.close()


    def get_external_list(self, file_url):
        fname, ext = os.path.splitext(file_url)

        has_macro = False
        err_code = 0

        if 'xlsm' in ext:
            has_macro = True

        try:
            wb = load_workbook(filename=file_url, data_only=True, keep_vba=has_macro, read_only=True)
            self.external_list.clear()
            self.external_file_url = file_url

            err_code += 1
            ws = wb['배정전정리']
            i = 0

            err_code += 1

            for row in ws.iter_rows(min_row=4):
                if row[0].value is None:
                    break

                if '미충원' in row[4].value:
                    type = '미충원'
                else:
                    type = '타시군전입'

                t = TeacherExternal(id=i, rank=row[0].value, type=type, region=row[1].value, position=row[3].value,
                                    school=row[2].value, name=row[4].value, birth=row[5].value, sex=row[6].value,
                                    major=row[8].value, career=row[7].value, first=row[9].value, second=row[10].value,
                                    third=row[11].value, ab_type=row[12].value, ab_start=row[13].value,
                                    ab_end=row[14].value, related_school=row[15].value, relation=row[16].value,
                                    relation_person=row[17].value, address=row[18].value, phone=row[19].value,
                                    email=row[21].value, vehicle=row[22].value, remarks=row[23].value)

                self.external_list.append(t)
                i += 1

            err_code += 1

            ws = wb['순위명부']

            for row in ws.iter_rows(min_row=3):
                if row[0].value is None:
                    break

                self.external_list[row[0].value-1].disposed = row[10].value

            err_code += 1

            ws = wb['시흥전출']
            i = 0

            for row in ws.iter_rows(min_row=2):
                if row[0].value is None:
                    break

                t = TeacherExternal(id=i, rank=0, type=row[0].value, region=row[1].value, position=None,
                                    school=row[4].value, name=row[2].value, birth=None, sex=None,
                                    major=None, career=None, first=None, second=None,
                                    third=None, ab_type=None, ab_start=None,
                                    ab_end=None, related_school=None, relation=None,
                                    relation_person=None, address=None, phone=None,
                                    email=None, vehicle=None, remarks=None)

                t.disposed = row[1].value
                self.gone_external.append(t)
                i += 1

        except TypeError as e:
            print(e)

            self.show_msg_box("{}시트 {}번째 행 서식을 확인해주세요.".format(ws.title, i + row[0].row), True)

            self.flag_external = False
            wb.close()

        except KeyError as e:
            print(e)

            if err_code == 1:
                err_sheet_name = '배정전정리'
            elif err_code == 2:
                err_sheet_name = '순위명부'
            elif err_code == 3:
                err_sheet_name = '시흥전출'

            self.show_msg_box("\'" + err_sheet_name + "\' 시트가 필요합니다.", True)
            self.flag_external = False
            wb.close()

        except Exception as e:
            print(e)

            if isinstance(e, InvalidFileException):
                self.show_msg_box("파일을 선택해주세요.", True)
            else:
                self.show_msg_box("올바른 파일을 선택해주세요.", True)

            if self.external_list.__len__() == 0:
                self.flag_external = False

        else:
            self.flag_external = True
            self.show_msg_box("성공적으로 불러왔습니다.", False)
            wb.close()

    def is_valid(self):
        if self.flag_internal and self.flag_external and self.flag_schools:
            return True

        else:
            msg = ''

            if not self.flag_schools:
                msg += '결충원 현황'

            if not self.flag_internal:
                if msg:
                    msg += ', '
                msg += '순위명부'

            if not self.flag_external:
                if msg:
                    msg += ', '
                msg += '타시군 전입명부'

            msg = '불러오지 않은 파일이 있습니다.\n' + msg

            self.show_msg_box(msg=msg, is_error=True)

            return False

    def start_program(self):
        self.post_thread.start()

    def show_next_view(self):
        print('internal {} school {} external {}'
              .format(self.internal_list.__len__(), self.school_list.__len__(), self.external_list.__len__()))

        self.field = WorkingField(internal=self.internal_list,
                                  external=self.external_list,
                                  schools=self.school_list,
                                  hash_school=self.hash_schools,
                                  designation=self.designation,
                                  gone=self.gone,
                                  unDeployed=self.unDeployed,
                                  recruit=self.recruit,
                                  vacancy=self.vacancy,
                                  controller=self, )

    def save(self, file_url):
        self.save_thread.result_file_url = file_url
        self.save_thread.start()

    def save_example(self, file, save_name):
        file_url = 'example_' + file + ".xlsx"
        wb = load_workbook(resource_path(file_url))

        wb.save(save_name + '.xlsx')

        self.show_msg_box('저장 완료', False)

    def update(self):
        self.update_thread.internal_file_url = self.internal_file_url
        self.update_thread.external_file_url = self.external_file_url
        self.update_thread.school_file_url = self.school_file_url
        self.update_thread.start()

    def show_msg_box(self, msg, is_error):
        msg_box = QMessageBox()

        if is_error:
            msg_box.setIcon(QMessageBox.Warning)
        else:
            msg_box.setIcon(QMessageBox.NoIcon)

        msg_box.setText(msg)
        msg_box.setWindowTitle("")
        msg_box.setStandardButtons(QMessageBox.Ok)
        result = msg_box.exec()

    def show_dialog_exit(self):
        msg_box = QMessageBox()
        msg_box.setWindowTitle("")
        msg_box.setText('종료하시겠습니까?')
        msg_box.setIcon(QMessageBox.Question)
        msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.No)
        result = msg_box.exec_()

        if result == QMessageBox.Ok:
            QApplication.quit()


def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)