
import openpyxl
from PyQt5 import QtCore
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break

from teacher_external import TeacherExternal

SLEEP = 1
PAGE_BREAK = 26


class SavingThread(QtCore.QThread):
    show_msg_box = QtCore.pyqtSignal(str, bool)
    set_state_result = QtCore.pyqtSignal(int)

    def __init__(self, **kwargs):
        super().__init__()
        self.internal = kwargs['internal']
        self.external = kwargs['external']
        self.schools = kwargs['schools']
        self.invited = kwargs['invited']
        self.priority = kwargs['priority']
        self.designation = kwargs['designation']
        self.gone = kwargs['gone']
        self.result_file_url = None
        self.controller = kwargs['controller']

        self.is_error = False
        self.msg = ''
        # self.finished.connect(self.controller.show_msg_box)

    def run(self) -> None:
        self.save()

    def save(self):
        self.is_error = False
        self.msg = ''

        self.make_result_file()

        if not self.is_error:
            self.msg = "저장되었습니다."
        self.show_msg_box.emit(self.msg, self.is_error)

    def make_result_file(self):
        try:
            wb = openpyxl.Workbook()
            self.counter = 0
            sheet1 = wb.active
            sheet1.title = '관내발령결과'
            self.make_result_sheet(sheet1, self.internal, '경기도시흥교육지원청 초등교사 정기인사(관내)-현임교순')

            sheet2 = wb.create_sheet('타시군(도)발령결과')
            self.make_result_sheet(sheet2, self.external, '경기도시흥교육지원청 초등교사 정기인사(관외)- 성명순')

            sheet3 = wb.create_sheet('초빙교사발령결과')
            self.make_result_sheet(sheet3, self.invited, '경기도시흥교육지원청 초등초빙교사 정기인사-성명순')

            sheet4 = wb.create_sheet('data-in')
            self.make_data_in(sheet4)

            sheet5 = wb.create_sheet('봉투라벨')
            self.make_pack_label(sheet5)

            wb.save(self.result_file_url)

            print('result saved')

        except Exception as e:
            print(e)
            self.is_error = True
            if self.msg:
                self.msg += ', '
            self.msg += '발령결과'
            wb.close()

    def make_result_sheet(self, sheet, teachers, title):
        fontStyle = Font(size="14", bold=True)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        sheet.merge_cells('A1:F1')
        title_cell = sheet['A1']
        title_cell.value = title
        title_cell.font = fontStyle
        title_cell.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        fontStyle = Font(size="10")

        sheet['A2'] = '유형'
        sheet['A2'].font = fontStyle
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet['A2'].border = thin_border

        sheet['B2'] = '지역'
        sheet['B2'].font = fontStyle
        sheet['B2'].alignment = Alignment(horizontal='center')
        sheet['B2'].border = thin_border

        sheet['C2'] = '소속교'
        sheet['C2'].font = fontStyle
        sheet['C2'].alignment = Alignment(horizontal='center')
        sheet['C2'].border = thin_border

        sheet['D2'] = '성명'
        sheet['D2'].font = fontStyle
        sheet['D2'].alignment = Alignment(horizontal='center')
        sheet['D2'].border = thin_border

        sheet['E2'] = '성별'
        sheet['E2'].font = fontStyle
        sheet['E2'].alignment = Alignment(horizontal='center')
        sheet['E2'].border = thin_border

        sheet['F2'] = '발령사항'
        sheet['F2'].font = fontStyle
        sheet['F2'].alignment = Alignment(horizontal='center')
        sheet['F2'].border = thin_border

        sheet.column_dimensions['E'].hidden = True
        sheet.column_dimensions['F'].width = 25.0

        row = 3
        for teacher in teachers:
            self.counter += 1
            self.set_state_result.emit(self.counter)
            if teacher.disposed is None:
                continue

            if isinstance(teacher, TeacherExternal):
                value = '타시군전입'
            else:
                value = '관내'

            a = sheet.cell(row=row, column=1, value=value)
            a.alignment = Alignment(horizontal='center')
            a.font = fontStyle

            if isinstance(teacher, TeacherExternal):
                value = teacher.region
            else:
                value = '시흥'
            b = sheet.cell(row=row, column=2, value=value)
            b.alignment = Alignment(horizontal='center')
            b.font = fontStyle

            c = sheet.cell(row=row, column=3, value=teacher.school)
            c.alignment = Alignment(horizontal='center')
            c.font = fontStyle

            print('aaaaaa')
            d = sheet.cell(row=row, column=4, value=teacher.name)
            d.alignment = Alignment(horizontal='center')
            d.font = fontStyle

            e = sheet.cell(row=row, column=5, value=teacher.sex)
            e.alignment = Alignment(horizontal='center')
            e.font = fontStyle

            f = sheet.cell(row=row, column=6, value=teacher.disposed.name + '등학교 근무를 명함')
            f.font = fontStyle

            row += 1

        g = sheet.cell(row=row, column=6
                       , value='이상 {}명\n경기도시흥교육지원청 교육장.\n끝.'.format(row-3))
        g.alignment = Alignment(wrap_text=True)
        g.font = fontStyle

    def make_data_in(self, sheet):
        # sheet.print_title_cols = 'A:H'  # the first two cols
        # sheet.print_title_rows = '1:1'  # the first row

        sheet.auto_filter.ref = 'A:H'

        fontStyle = Font(size="11")
        alignment = Alignment(horizontal='center')

        self.write_to_cell(sheet['A1'], '순', fontStyle, alignment)
        self.write_to_cell(sheet['B1'], '유형', fontStyle, alignment)
        self.write_to_cell(sheet['C1'], '소속지', fontStyle, alignment)
        self.write_to_cell(sheet['D1'], '소속교', fontStyle, alignment)
        self.write_to_cell(sheet['E1'], '성명', fontStyle, alignment)
        self.write_to_cell(sheet['F1'], '임지', fontStyle, alignment)
        self.write_to_cell(sheet['G1'], '성별', fontStyle, alignment)
        self.write_to_cell(sheet['H1'], '생년월일', fontStyle, alignment)

        i = 1
        row = 2

        for teacher in self.internal:
            self.counter += 1
            self.set_state_result.emit(self.counter)
            if teacher.disposed is None:
                continue
            print('bbbbb')
            self.write_to_cell(sheet.cell(row=row, column=1), i, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=2), '관내', fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=3), '시흥', fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=4), teacher.school, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=5), teacher.name, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=6), teacher.disposed.name, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=7), teacher.sex, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=8), teacher.birth, fontStyle, alignment)

            i += 1
            row += 1

        for teacher in self.invited:
            self.counter += 1
            self.set_state_result.emit(self.counter)
            if teacher.disposed is None:
                continue
            print('ccccc')
            self.write_to_cell(sheet.cell(row=row, column=1), i, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=2), '초빙', fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=3), '시흥', fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=4), teacher.school, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=5), teacher.name, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=6), teacher.disposed.name, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=7), teacher.sex, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=8), teacher.birth, fontStyle, alignment)

            i += 1
            row += 1

        for teacher in self.external:
            self.counter += 1
            self.set_state_result.emit(self.counter)
            if teacher.disposed is None:
                continue

            self.write_to_cell(sheet.cell(row=row, column=1), i, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=2), '타시군전입', fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=3), teacher.region, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=4), teacher.school, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=5), teacher.name, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=6), teacher.disposed.name, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=7), teacher.sex, fontStyle, alignment)
            self.write_to_cell(sheet.cell(row=row, column=8), teacher.birth, fontStyle, alignment)

            i += 1
            row += 1

    def write_to_cell(self, cell, value, font, alignment, color=PatternFill(), border=None):
        cell.value = value
        cell.font = font
        cell.alignment = alignment
        cell.fill = color
        cell.border = border

    def make_pack_label(self, sheet):
        fontStyle = Font(size="11")
        alignment = Alignment(horizontal='center')

        sheet.column_dimensions['A'].width = 12.00
        sheet.column_dimensions['B'].width = 10.33
        sheet.column_dimensions['C'].width = 10.67
        sheet.column_dimensions['D'].width = 13.33
        sheet.column_dimensions['E'].width = 8.17
        sheet.column_dimensions['F'].width = 8.17
        sheet.column_dimensions['G'].width = 8.17
        sheet.column_dimensions['H'].width = 0.1
        sheet.column_dimensions['I'].width = 9.33
        sheet.column_dimensions['J'].width = 13.67
        sheet.column_dimensions['K'].width = 9.17
        sheet.column_dimensions['L'].width = 13.33
        sheet.page_setup.fitToWidth = True
        print_area = 'A{}:L{}'.format(1, self.schools.__len__()*25 + 3)
        sheet.print_area = print_area
        sheet.sheet_view.view = 'pageBreakPreview'
        sheet.set_printer_settings(sheet.PAPERSIZE_A4, sheet.ORIENTATION_LANDSCAPE)
        sheet.page_margins = PageMargins(left=0.454722225666046,
                                         right=0.312083333730698,
                                         top=1.30750000476837,
                                         bottom=0.75,
                                         header=0.698055565357208,
                                         footer=0.314861118793488)

        for i in range(0, self.schools.__len__()):
            row = i * PAGE_BREAK

            row_break = Break(id=row)  # create Break obj
            sheet.row_breaks.append(row_break)
            self.make_label_header(sheet, row)

            for teacher in self.designation[i]:

                self.write_to_cell(sheet.cell(row=row+3, column=1), teacher.disposed.name, fontStyle, alignment)
                if isinstance(teacher, TeacherExternal):
                    if '미충원' in teacher.name:
                        self.write_to_cell(sheet.cell(row=row+3, column=2), '미충원', fontStyle, alignment)
                    else:
                        self.write_to_cell(sheet.cell(row=row+3, column=2), '타시군', fontStyle, alignment)
                    self.write_to_cell(sheet.cell(row=row+3, column=3), teacher.region, fontStyle, alignment)

                else:
                    self.write_to_cell(sheet.cell(row=row+3, column=2), '관내', fontStyle, alignment)
                    self.write_to_cell(sheet.cell(row=row+3, column=3), '시흥', fontStyle, alignment)

                self.write_to_cell(sheet.cell(row=row+3, column=4), teacher.school, fontStyle, alignment)
                self.write_to_cell(sheet.cell(row=row+3, column=5), teacher.name, fontStyle, alignment)
                self.write_to_cell(sheet.cell(row=row+3, column=6), teacher.birth, fontStyle, alignment)
                self.write_to_cell(sheet.cell(row=row+3, column=7), teacher.sex, fontStyle, alignment)
                row += 1
                # print(self.designation[j])
                self.counter += 1
                self.set_state_result.emit(self.counter)

            row = i * PAGE_BREAK
            for teacher in self.gone[i]:
                self.counter += 1
                self.set_state_result.emit(self.counter)
                if '타시군' in teacher.type:
                    continue
                    # self.write_to_cell(sheet.cell(row=row, column=9), '타시군', fontStyle, alignment)
                    # self.write_to_cell(sheet.cell(row=row, column=9), teacher.region, fontStyle, alignment)

                else:
                    self.write_to_cell(sheet.cell(row=row+3, column=9), '관내', fontStyle, alignment)
                    # self.write_to_cell(sheet.cell(row=row, column=3), '시흥', fontStyle, alignment)
                self.write_to_cell(sheet.cell(row=row+3, column=10), teacher.school, fontStyle, alignment)
                self.write_to_cell(sheet.cell(row=row+3, column=11), teacher.name, fontStyle, alignment)
                self.write_to_cell(sheet.cell(row=row+3, column=12), teacher.disposed.name, fontStyle, alignment)
                row += 1

                # print(self.gone[j])

    def make_label_header(self, sheet, row):
        fontStyle = Font(size="11")
        alignment = Alignment(horizontal='center')
        fill_coming = PatternFill(patternType='solid', fgColor=Color('5182BB'))
        fill_gone = PatternFill(patternType='solid', fgColor=Color('D59493'))
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        sheet.cell(row=row+1, column=1).value = '전입'
        sheet.cell(row=row+1, column=1).font = fontStyle
        sheet.cell(row=row+1, column=1).alignment = alignment
        sheet.cell(row=row+1, column=1).fill = fill_coming
        sheet.cell(row=row+1, column=1).border = thin_border

        sheet.merge_cells(start_row=row+1, end_row=row+1, start_column=1, end_column=7)

        # sheet.merge_cells('')
        sheet.cell(row=row+1, column=9).value = '전출'
        sheet.cell(row=row+1, column=9).font = fontStyle
        sheet.cell(row=row+1, column=9).alignment = alignment
        sheet.cell(row=row+1, column=9).fill = fill_gone
        sheet.cell(row=row+1, column=9).border = thin_border

        sheet.merge_cells(start_row=row+1, end_row=row+1, start_column=9, end_column=12)

        self.write_to_cell(sheet.cell(row=row+2, column=1), '신임교', fontStyle, alignment, fill_coming, thin_border)
        self.write_to_cell(sheet.cell(row=row+2, column=2), '전보유형', fontStyle, alignment, fill_coming, thin_border)
        self.write_to_cell(sheet.cell(row=row+2, column=3), '소속청', fontStyle, alignment, fill_coming, thin_border)
        self.write_to_cell(sheet.cell(row=row+2, column=4), '소속교', fontStyle, alignment, fill_coming, thin_border)
        self.write_to_cell(sheet.cell(row=row+2, column=5), '성명', fontStyle, alignment, fill_coming, thin_border)
        self.write_to_cell(sheet.cell(row=row+2, column=6), '생년월일', fontStyle, alignment, fill_coming, thin_border)
        self.write_to_cell(sheet.cell(row=row+2, column=7), '성별', fontStyle, alignment, fill_coming, thin_border)

        self.write_to_cell(sheet.cell(row=row+2, column=9), '전보유형', fontStyle, alignment, fill_gone, thin_border)
        self.write_to_cell(sheet.cell(row=row+2, column=10), '소속', fontStyle, alignment, fill_gone, thin_border)
        self.write_to_cell(sheet.cell(row=row+2, column=11), '성명', fontStyle, alignment, fill_gone, thin_border)
        self.write_to_cell(sheet.cell(row=row+2, column=12), '임지지정', fontStyle, alignment, fill_gone, thin_border)
